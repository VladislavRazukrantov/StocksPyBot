from datetime import datetime

import telebot
import finnhub
import openpyxl
from telebot import types

bot = telebot.TeleBot('5707455690:AAH_dra68o20IGv0q94Bz7O1m3mHdoS73dw')
finnhub_client = finnhub.Client(api_key="cdvig5iad3i78d1oq5j0cdvig5iad3i78d1oq5jg")
keyboard = types.InlineKeyboardMarkup()

name_of_stock = ''
amount_of_stocks = 0

title = ''
amount = 0
price = 0

name_column = 1 #имя акции, количество и цена - даем соответсвующие номера столбцов
amount_column = 2
price_column = 3

list_of_symbols = []

def init(id):
    wb = openpyxl.load_workbook('usr.xlsx') #получаем объект рабочей книги
    wb.create_sheet(title=str(id))  #создаем лист
    sh = wb[str(id)]
    current_row_cell = sh.cell(row=1, column=1) #доступ к ячейке
    current_profit_cell = sh.cell(row=1, column=2) #доступ к ячейке, тут будет прибыль общая
    current_row_cell.value = 2   #присваиваем значение
    current_profit_cell.value = 0 #изначально прибыли нет
    wb.save('usr.xlsx') #сохраняем


@bot.message_handler(commands=['start', 'help']) #реагируем на входящие сообщения
def start(message):
    #message - json объект, хранящий информацию об отправителе, чате и сообщении
    init(message.from_user.id)
    start_message = 'Приветствую! Это финансовый бот, который позволит Вам упростить работу с акциями!\n\n' + \
                    'Команды:\n\n' + \
                    '/portfolio - Посмотреть портфель\n' + \
                    '/quotations - Посмотреть котировки \n'
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True) #пользовательская клавиатура
    key_portfolio = types.KeyboardButton(text='Портфель') #добавляем кнопки
    key_quotations = types.KeyboardButton(text='Котировки')
    markup.add(key_portfolio, key_quotations)
    bot.send_message(message.chat.id, start_message, reply_markup=markup)  #третий аргумент - меняем клавиатуру на кнопочную


@bot.message_handler(commands=['back'])
def back(message):
    start_message = 'Приветствую! Это финансовый бот, который позволит Вам упростить работу с акциями!\n\n' + \
                    'Команды:\n\n' + \
                    '/portfolio - Посмотреть портфель\n' + \
                    '/quotations - Посмотреть котировки \n'
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_portfolio = types.KeyboardButton(text='Портфель')
    key_quotations = types.KeyboardButton(text='Котировки')
    #key_graphics = types.KeyboardButton(text='Графики')
    markup.add(key_portfolio, key_quotations)
    bot.send_message(message.chat.id, start_message, reply_markup=markup)


@bot.message_handler(commands=['portfolio'])
def portfolio(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True) #пользовательская клавиатура
    key_show_stocks = types.KeyboardButton(text='Показать акции')  #кнопки
    key_new_stock = types.KeyboardButton(text='Купить акции')
    key_sell_stock = types.KeyboardButton(text='Продать акции')
    key_back_to_menu = types.KeyboardButton(text='Вернуться в меню')
    markup.add(key_show_stocks, key_new_stock, key_sell_stock, key_back_to_menu) #добавляем все кнопки
    bot.send_message(message.chat.id, "Выберите команду", reply_markup=markup)


@bot.message_handler(commands=['quotations'])
def quotations(message):
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_of_stock_for_quotations) #передаем сообщение и след шаг, после ответа пользователя


@bot.message_handler(content_types=['text']) #обработка текстового сообщения
def random_answers(message):
    if message.text == 'Портфель':
        portfolio(message)
    # if message.text == 'Графики':
    #     graphics(message)
    if message.text == 'Котировки':
        quotations(message)
    if message.text == 'Показать акции':
        show_stocks(message)
    if message.text == 'Купить акции':
        add_stock(message)
    if message.text == 'Продать акции':
        sell_stocks(message)
    if message.text == 'Вернуться в меню':
        back(message)


# ПОРТФОЛИО
def show_stocks(message): #Переходим при выборе кнопки показать акции
    global current_row #текущая строка
    global current_profit
    global name_column
    global price_column

    wb = openpyxl.load_workbook('usr.xlsx') #получаем нашу рабочую книгу
    sh = wb[str(message.from_user.id)]      #получаем страницу с нужным пользователем
    current_row_cell = sh.cell(row=1, column=1) #получаем значение ячейки 1:1
    current_profit_cell = sh.cell(row=1, column=2)

    current_row = current_row_cell.value
    current_profit = current_profit_cell.value
    if current_row == 2: #изначально у нас два
        answer = '💼 Портфель\n\nПусто\n\n'
        if (current_profit >0):
            answer += 'Прибыль за время инвестиций: ' + str(current_profit) +  ' $' + '\n'
        else:
            answer += 'Убыток за время инвестиций: ' + str(current_profit) + ' $' + '\n'
        bot.send_message(message.chat.id, answer)
    else:
        answer = '💼 Портфель💰💰💰 \n\n'
        for i in range(2, sh.max_row + 1):
            answer += str(i - 1) + ". "
            n = sh.cell(row=i, column=name_column) #номер в таблице - 1, мы это выше объявили
            a = sh.cell(row=i, column=amount_column) #получаем значение
            p = sh.cell(row=i, column=price_column)
            answer += (str(n.value) + ': ').ljust(6) + (str(a.value) + ' шт. цена: ').ljust(12) + str(p.value) + '$ \n\n'
        if (current_profit > 0):
            answer += 'Прибыль за время инвестиций: ' + str(current_profit) + ' $' + '\n'
        else:
            answer += 'Убыток за время инвестиций: ' + str(current_profit) + ' $' + '\n'
        answer += '\n'
        bot.send_message(message.chat.id, answer)


def add_stock(message): #попадаем когда пользователь ввел купить акции
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_buy) #переходим дальше


def sell_stocks(message): #попадаем когда пользователь ввел продать акции
    bot.send_message(message.chat.id, 'Введите название акции')
    bot.register_next_step_handler(message, get_title_sell)
    # проверка на существуемость: если нет то ошибка и еще раз


# получение данных для покупки
def get_title_buy(message):
    stocks_variants = finnhub_client.symbol_lookup(message.text)  #ищем подходящие варианты после ввода пользователя
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций. Попробуйте еще раз!')
        #bot.register_next_step_handler(message, get_title_buy)
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer) #выводим список найденных акций и переходим дальше по обработчику
        bot.register_next_step_handler(message, get_title_buy_next)


def get_title_buy_next(message):
    global title #для имени акции
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Указанная акция не найдена')
        #bot.register_next_step_handler(message, get_title_buy_next)
    else:
        title = message.text
        list_of_symbols.clear()
        bot.send_message(message.chat.id, 'Введите количество акций')
        bot.register_next_step_handler(message, get_amount_buy)


def get_amount_buy(message):
    global amount #для количества
    try:
        amount = int(message.text) #количество, которое ввел пользователь
        bot.send_message(message.chat.id, 'Введите цену покупки\nПримечание: цена - число')
        bot.register_next_step_handler(message, get_price_buy)
    except: #обрабатываем ввод пользователя (он вводит количество акций)
        bot.send_message(message.chat.id, 'Введите количество акций\nПримечание: количество - целое число')
        bot.register_next_step_handler(message, get_amount_buy)


def get_price_buy(message): #пользователь вводит цену покупки и попадает сюда
    global price
    try:
        price = int(message.text)
        buy(message)     #что нужно чтобы произошла покупка
    except:
        bot.send_message(message.chat.id, 'Введите цену акции\nПримечание: цена - число')
        bot.register_next_step_handler(message, get_price_buy)


# попадаем после ввода названия акции
def get_title_sell(message):
    global title #для названия акции
    title = message.text
    bot.send_message(message.chat.id, 'Введите количество акций\nПримечание: количество - целое число')
    bot.register_next_step_handler(message, get_amount_sell)


def get_amount_sell(message): #попали после ввода пользователем количества
    global amount
    try:
        amount = int(message.text) #получаем число
        bot.send_message(message.chat.id, 'Введите цену продажи\nПримечание: цена - число')
        bot.register_next_step_handler(message, get_price_sell)
    except: # если было введено не число, то еще раз просим ввести
        bot.send_message(message.chat.id, 'Введите количество акций\nПримечание: количество - целое число')
        bot.register_next_step_handler(message, get_amount_sell)


def get_price_sell(message): # после того, как ввели цену
    global price
    try:
        price = int(message.text)
        sell(message) #продажа, что при этом происходит
    except:
        bot.send_message(message.chat.id, 'Введите цену акции\nПримечание: цена - число')
        bot.register_next_step_handler(message, get_price_sell)


# покупка и продажа под капотом
def buy(message):
    global title #нужно объявить иначе не сможем использовать
    global amount
    global price
    global current_row #текущая строка

    global name_column
    global price_column
    global amount_column

    if amount <= 0 or price <= 0:  #то что ввел пользователь
        answer = '💼 Дорогой инвестор!\n\n '
        answer += 'Количество акций и цена должны быть положительными величинами. '
        bot.send_message(message.chat.id, answer)
        show_stocks(message) # выводим портфель
        return

    wb = openpyxl.load_workbook('usr.xlsx')
    sh = wb[str(message.from_user.id)]  #получаем страницу для пользователя

    current_row_cell = sh.cell(row=1, column=1)
    current_row = current_row_cell.value #смотрим значение ячейки 1:1

    alreadyExist, row_value_existed = is_existed(sh) #существует или нет, возвращается 2 значения: существует ли акция
    # и строка в которой найдена акция

    if not alreadyExist: # если акции до этого не было
        current_row_cell = sh.cell(row=1, column=1) #получаем первую ячейку, в 1:1 та строка, куда записывать будем
        t = sh.cell(row=current_row, column=name_column) #получаем доступ к ячейкам
        a = sh.cell(row=current_row, column=amount_column)
        p = sh.cell(row=current_row, column=price_column)
        t.value = title #присваиваем им соответсвующие значения того, что выбрал пользователь
        a.value = amount
        p.value = price
        current_row_cell.value = current_row + 1 #увеличиваем число в 1:1 на единицу

    else: # если акция есть
        a = sh.cell(row=row_value_existed, column=amount_column) #тут получаем количество и цену, название не надо
        p = sh.cell(row=row_value_existed, column=price_column)
        tmp = int(a.value)
        a.value = tmp + int(amount) #количество, которое было + то, что ввел пользователь
        p.value = (int(p.value) * tmp + int(price) * int(amount)) / (tmp + int(amount))
    #новая цена = (старая цена * старое количество + цена * кол-во, которое ввел пользователь) / (старое + новое количество)
    wb.save('usr.xlsx')  # все сохраняем
    title = ''
    amount = 0
    price = 0
    show_stocks(message) # выводим портфель


def sell(message): #продажа
    global title
    global amount
    global price
    global current_profit

    global name_column
    global price_column
    global amount_column

    if amount <= 0 or price <= 0:
        answer = '💼 Дорогой инвестор!\n\n '
        answer += ' Количество акций и цена должны быть положительными величинами. '
        bot.send_message(message.chat.id, answer)
        show_stocks(message)
        return

    wb = openpyxl.load_workbook('usr.xlsx') #получаем нашу рабочую книгу
    sh = wb[str(message.from_user.id)] #обращаемся к нужной странице
    current_profit_cell = sh.cell(row=1, column=2)
    current_profit = current_profit_cell.value #ячейка с прибылью
    alreadyExist, row_value_existed = is_existed(sh) #проверяем есть ли акция и если есть, то какая строка

    if alreadyExist: #если есть
        a = sh.cell(row=row_value_existed, column=amount_column) #получаем доступ к ячейкам данной акции
        p = sh.cell(row=row_value_existed, column=price_column)
        t = sh.cell(row=row_value_existed, column=name_column)
        tmp = a.value
        if a.value < int(amount): #сравниваем количество с тем, что ввел пользователь
            answer = '💼 Дорогой инвестор!\n\n '
            answer += ' Количество акций, которые хотите продать, не может быть больше, чем есть у Вас в портфеле. '
            bot.send_message(message.chat.id, answer)
            show_stocks(message)
            return
        a.value = tmp - int(amount) #новое количество = старое - введенное пользователем
        print(a.value)
        if a.value == 0:  #убираем строку в листе и указываем какую строку
            wb = openpyxl.load_workbook('usr.xlsx')  # получаем нашу рабочую книгу
            # print(id_user)
            sh = wb[str(message.from_user.id)]  # получаем нужный лист по id
            if row_value_existed == sh.max_row:  # если строка, которую надо занулить последняя
                m = sh.cell(row=row_value_existed, column=name_column)  # получаем значения ячеек
                b = sh.cell(row=row_value_existed, column=amount_column)
                l = sh.cell(row=row_value_existed, column=price_column)
                print(row_value_existed)
                m.value = None  # зануляем их
                b.value = None
                l.value = None
                # print(a.value, n.value, p.value)
                current_row_cell = sh.cell(row=1, column=1)
                current_row_cell.value = int(current_row_cell.value) - 1  # уменьшаем на 1 количество строк в 1:1
                wb.save('usr.xlsx')
            else:# с ячейками, которые выше удаляемой мы ничего не делаем, а вот ниже удаляемой строки мы берем
                m = sh.cell(row=row_value_existed + 1, column=name_column)  # получаем доступ к ячейкам следующей строки
                b = sh.cell(row=row_value_existed + 1, column=amount_column)
                l = sh.cell(row=row_value_existed + 1, column=price_column)

                column_of_name = [m.value]  # списки, в которых значения ячеек
                column_of_amount = [b.value]
                column_of_price = [l.value]

                for i in range(row_value_existed + 2, sh.max_row + 1):
                    m = sh.cell(row=i, column=name_column)  # получаем доступ к след ячейкам
                    b = sh.cell(row=i, column=amount_column)
                    l = sh.cell(row=i, column=price_column)

                    column_of_name.append(m.value)  # вставка их значений в конец
                    column_of_amount.append(b.value)
                    column_of_price.append(l.value)

                column_of_name.reverse()  # меняем порядок элементов в списке на противоположный
                column_of_amount.reverse()
                column_of_price.reverse()

                for i in range(row_value_existed, sh.max_row):  # начиная от строки, которую удаляем до последней
                    m = sh.cell(row=i,
                                column=name_column)  # получаем доступ сначала к строке, которую удаляем, потом ниже и т.д.
                    b = sh.cell(row=i, column=amount_column)
                    l = sh.cell(row=i, column=price_column)
                    m.value = column_of_name.pop()  # для этой строки вставляем значение последнего элемента (для этого список переворачивали)
                    b.value = column_of_amount.pop()  # то есть строку которую удаляем заменяем на следующую и так далее
                    l.value = column_of_price.pop()

                m = sh.cell(row=sh.max_row,
                            column=name_column)  # последнюю строку зануляем, поскольку все вверх на 1 перешло
                b = sh.cell(row=sh.max_row, column=amount_column)
                l = sh.cell(row=sh.max_row, column=price_column)
                m.value = None
                b.value = None
                l.value = None

                current_row_cell = sh.cell(row=1, column=1)
                current_row_cell.value = int(current_row_cell.value) - 1  # уменьшаем количество строк
                wb.save('usr.xlsx')
        else:
            wb.save('usr.xlsx')

        answer = '💼 Время - деньги\n\n'
        answer += 'Вы продали акции ' + t.value + ' в количестве ' + str(amount) + ' \n\n'
        if a.value > 0:
            answer += 'Осталось акций компании ' + t.value + ' в количестве ' + str(a.value) + ' \n\n'
        profit = int(price) * int(amount) - int(p.value) * int(amount)
        current_profit_cell.value += profit
        if profit > 0:
            answer += 'Поздравляем с успешной сделкой! Прибыль от сделки: ' + str(profit) + ' $' + '\n'
        else:
            answer += 'Убыток от сделки: ' + str(profit) + ' $' + '\n'
        if (current_profit_cell.value > 0):
            answer += 'Прибыль за время инвестиций: ' + str(current_profit_cell.value) +  ' $'+ '\n'
        else:
            answer += 'Убыток за время инвестиций: ' + str(current_profit_cell.value) + ' $' +'\n'
        bot.send_message(message.chat.id, answer)
    else:
        answer = '💼 В портфеле\n\nНет акций с титром ' + title
        bot.send_message(message.chat.id, answer)
    wb.save('usr.xlsx')
    title = ''
    amount = 0
    price = 0


def is_existed(sh): #страница с пользователем
    current_row_cell = sh.cell(row=1, column=1) #получаем значение ячейки
    max_row = current_row_cell.value #это сколько строк у нас будет
    alreadyExist = False

    cntr = 1
    for row in sh.iter_rows(max_row=max_row): #итерируемся по строкам
        for cell in row: #cell - ячейка
            if str(cell.value) == title: #сравниваем значение ячейки с именем акции
                alreadyExist = True  #уже эта акция есть
                break
        if alreadyExist: #если акция существует количество увеличивать не надо
            break
        cntr += 1 #номер строки, где акция находится
    return alreadyExist, cntr

# КОТИРОВКИ
def get_title_of_stock_for_quotations(message): #после введения пользователем акции идем сюда
    global list_of_symbols
    stocks_variants = finnhub_client.symbol_lookup(message.text) #находятся наиболее подходящие акции по вводу пользователя
    if stocks_variants['count'] == 0:
        bot.send_message(message.chat.id, 'Не найдено подходящих акций')
    else:
        answer = 'Мы нашли похожие акции, выберите одну из них:\n'
        for stockEx in stocks_variants["result"]:
            answer += ('• ' + stockEx["symbol"] + '(' + stockEx["description"] + ', ' + stockEx["type"] + ')' + '\n')
            list_of_symbols.append(stockEx["symbol"])
        bot.send_message(message.chat.id, answer)
        bot.register_next_step_handler(message, get_title_of_stock_for_quotations_next) #после ввода пользователя идем дальше


def get_title_of_stock_for_quotations_next(message):
    global list_of_symbols
    if message.text not in list_of_symbols:
        bot.send_message(message.chat.id, 'Указанная акция не найдена')
        #bot.register_next_step_handler(message, get_title_of_stock_for_quotations_next)
    else:
        prices = finnhub_client.quote(message.text) #получаем данные о котировке
        answer = 'Текущая цена: ' + str(prices["c"]) + ' $ \n' \
                                                       'Наивысшая цена сегодня: ' + str(prices["h"]) + ' $ \n' \
                                                                                                       'Наименьшая цена сегодня: ' + str(
            prices["l"]) + ' $ \n' \
                           'Время: ' + str(datetime.fromtimestamp(prices["t"])) + '\n'

        bot.send_message(message.chat.id, answer)
        list_of_symbols.clear()

bot.polling(none_stop=True)
